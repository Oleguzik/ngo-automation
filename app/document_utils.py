"""
Universal document text extraction utility.

Supports multiple file formats for Phase 3 MVP:
- PDF (PyPDF2)
- XLSX/Excel (openpyxl)
- CSV (pandas)
- Images (Pillow + pytesseract OCR)

Reference: docs/00-spec-phase4.md - Document Processing Requirements
           docs/02-architecture-phase4.md - Text Extraction Pipeline
"""

from PyPDF2 import PdfReader
from PIL import Image
import pytesseract
import pandas as pd
import openpyxl
from typing import Optional, Tuple
import io
import logging

logger = logging.getLogger(__name__)


def extract_text_from_pdf(file_bytes: bytes) -> Optional[str]:
    """
    Extract text content from PDF file bytes.
    
    Uses PyPDF2 to extract text from all pages.
    Includes page markers for multi-page documents.
    
    Args:
        file_bytes: PDF file as bytes
        
    Returns:
        Extracted text string with page separators, or None if extraction failed
        
    Raises:
        Exception: Logged but not raised; returns None on failure
    """
    try:
        pdf_file = io.BytesIO(file_bytes)
        pdf_reader = PdfReader(pdf_file)
        
        text_parts = []
        for page_num, page in enumerate(pdf_reader.pages):
            page_text = page.extract_text()
            if page_text.strip():
                text_parts.append(f"--- Page {page_num + 1} ---\n{page_text}")
        
        full_text = "\n\n".join(text_parts)
        return full_text if full_text.strip() else None
        
    except Exception as e:
        logger.error(f"PDF extraction error: {str(e)}")
        return None


def extract_text_from_xlsx(file_bytes: bytes) -> Optional[str]:
    """
    Extract text content from Excel XLSX file.
    
    Reads all sheets and cells, formatting as pipe-separated values.
    Skips completely empty rows.
    Useful for bank statements, financial reports, donation tracking sheets.
    
    Args:
        file_bytes: XLSX file as bytes
        
    Returns:
        Extracted text with sheet names, headers, and data rows formatted as:
        --- Sheet: SheetName ---
        Header1 | Header2 | Header3
        Value1 | Value2 | Value3
        
        Returns None if extraction failed
    """
    try:
        excel_file = io.BytesIO(file_bytes)
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            
            # Extract all non-empty cells
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(val.strip() for val in row_values):
                    rows_data.append(" | ".join(row_values))
            
            text_parts.append("\n".join(rows_data))
        
        full_text = "\n\n".join(text_parts)
        return full_text if full_text.strip() else None
        
    except Exception as e:
        logger.error(f"XLSX extraction error: {str(e)}")
        return None


def extract_text_from_csv(file_bytes: bytes) -> Optional[str]:
    """
    Extract text content from CSV file.
    
    Auto-detects encoding (UTF-8, Latin-1, ISO-8859-1) for international support.
    Formats output as pipe-separated columns for consistency with XLSX.
    Useful for bank exports, donation records, expense logs.
    
    Args:
        file_bytes: CSV file as bytes
        
    Returns:
        Extracted text formatted as:
        --- CSV Data ---
        Header1 | Header2 | Header3
        ────────────────────────────
        Value1 | Value2 | Value3
        
        Returns None if extraction or encoding detection failed
    """
    try:
        csv_file = io.BytesIO(file_bytes)
        
        # Try different encodings for international character support
        for encoding in ['utf-8', 'latin-1', 'iso-8859-1']:
            try:
                csv_file.seek(0)
                df = pd.read_csv(csv_file, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            logger.error("Could not decode CSV with any supported encoding")
            return None
        
        # Convert DataFrame to text representation
        text_parts = []
        
        # Add header
        text_parts.append("--- CSV Data ---")
        text_parts.append(" | ".join(df.columns.astype(str)))
        text_parts.append("-" * 80)
        
        # Add rows
        for _, row in df.iterrows():
            row_text = " | ".join(row.astype(str))
            text_parts.append(row_text)
        
        full_text = "\n".join(text_parts)
        return full_text if full_text.strip() else None
        
    except Exception as e:
        logger.error(f"CSV extraction error: {str(e)}")
        return None


def extract_text_from_image(file_bytes: bytes) -> Optional[str]:
    """
    Extract text from image using Tesseract OCR.
    
    Supports: PNG, JPG, JPEG, GIF, BMP, TIFF
    
    Uses pytesseract (Python wrapper for Tesseract OCR engine).
    Automatically converts images to RGB if needed.
    Useful for scanned receipts, invoices, donation letters.
    
    Note: Requires tesseract-ocr system package installed (see Dockerfile)
    OCR accuracy depends on image quality and resolution (300+ DPI recommended)
    
    Args:
        file_bytes: Image file as bytes
        
    Returns:
        Extracted text via OCR, or None if extraction failed
        
    Raises:
        Exception: Logged but not raised; returns None on failure
    """
    try:
        image_file = io.BytesIO(file_bytes)
        image = Image.open(image_file)
        
        # Convert to RGB if necessary (for PNG with transparency, etc.)
        if image.mode not in ('RGB', 'L'):
            image = image.convert('RGB')
        
        # Perform OCR with English and German language support
        text = pytesseract.image_to_string(image, lang='eng+deu')
        
        return text.strip() if text.strip() else None
        
    except Exception as e:
        logger.error(f"Image OCR error: {str(e)}")
        return None


def extract_text_from_file(file_bytes: bytes, file_type: str, filename: str) -> Tuple[Optional[str], str]:
    """
    Universal text extraction function that routes to appropriate extractor.
    
    Intelligently detects file type from MIME type and/or filename extension.
    Routes to PDF, XLSX, CSV, or Image extractor accordingly.
    
    Args:
        file_bytes: File content as bytes
        file_type: MIME type (e.g., 'application/pdf', 'image/png')
        filename: Original filename for fallback extension detection
        
    Returns:
        Tuple of (extracted_text: Optional[str], file_format: str)
        file_format is one of: 'pdf', 'xlsx', 'csv', 'image', 'unsupported'
        
    Example:
        >>> text, fmt = extract_text_from_file(file_bytes, 'application/pdf', 'invoice.pdf')
        >>> if fmt == 'pdf':
        ...     print(f"Extracted {len(text)} chars from PDF")
    """
    # Normalize file type
    file_type_lower = file_type.lower()
    filename_lower = filename.lower()
    
    # PDF
    if 'pdf' in file_type_lower or filename_lower.endswith('.pdf'):
        text = extract_text_from_pdf(file_bytes)
        return (text, "pdf")
    
    # Excel XLSX (also handles .xls if openpyxl supports it)
    elif 'spreadsheet' in file_type_lower or 'excel' in file_type_lower or filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        text = extract_text_from_xlsx(file_bytes)
        return (text, "xlsx")
    
    # CSV
    elif 'csv' in file_type_lower or filename_lower.endswith('.csv'):
        text = extract_text_from_csv(file_bytes)
        return (text, "csv")
    
    # Images (PNG, JPG, JPEG, GIF, BMP, TIFF)
    elif any(img_type in file_type_lower for img_type in ['image/', 'png', 'jpeg', 'jpg', 'gif', 'bmp', 'tiff']) \
         or any(filename_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif']):
        text = extract_text_from_image(file_bytes)
        return (text, "image")
    
    else:
        logger.warning(f"Unsupported file type: {file_type} ({filename})")
        return (None, "unsupported")


def get_supported_file_types() -> dict:
    """
    Return dictionary of supported file types and their descriptions.
    
    Used for API validation and documentation.
    
    Returns:
        Dict structure:
        {
            "pdf": {
                "extensions": [".pdf"],
                "mime_types": ["application/pdf", ...],
                "description": "PDF documents"
            },
            ...
        }
    """
    return {
        "pdf": {
            "extensions": [".pdf"],
            "mime_types": ["application/pdf", "application/x-pdf"],
            "description": "PDF documents (invoices, receipts, statements)"
        },
        "excel": {
            "extensions": [".xlsx", ".xls"],
            "mime_types": [
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel"
            ],
            "description": "Excel spreadsheets (financial reports, donation tracking)"
        },
        "csv": {
            "extensions": [".csv"],
            "mime_types": ["text/csv", "application/csv"],
            "description": "CSV files (bank exports, expense logs)"
        },
        "images": {
            "extensions": [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif"],
            "mime_types": ["image/png", "image/jpeg", "image/gif", "image/bmp", "image/tiff"],
            "description": "Image files (scanned documents, photos) - OCR extraction"
        }
    }


if __name__ == "__main__":
    """Test utility - display supported file types"""
    print("📄 Document Extraction Utility - Supported Formats\n")
    
    supported = get_supported_file_types()
    for category, info in supported.items():
        ext_str = ", ".join(info['extensions'])
        print(f"✅ {category.upper():<8} {ext_str:<20} - {info['description']}")
