"""
Excel workbook scaffolding for Phase 4 (Sprint 3, Day 11).

This module introduces the foundational class for GoBD-compliant Excel
report generation without touching contracts in other modules. It focuses
on workbook initialization, German locale-friendly formats, global styles,
and deterministic filename conventions. Subsequent days will add sheet
creation and data mapping.

Dependencies: openpyxl (expected to be available in the environment).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, Sequence
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class GermanExcelFormats:
    """Common German-friendly number/date formats used across the workbook.

    Notes:
    - Excel number formats are locale-agnostic strings; we choose codes that
      display comma as decimal separator and period as thousands separator
      where supported by the user's Excel locale settings.
    - Currency format uses the Euro symbol and two decimals.
    - Date format follows German convention DD.MM.YYYY.
    """

    date: str = "DD.MM.YYYY"
    # Currency with thousands separator and comma decimals; include Euro symbol.
    # Excel may interpret locale-specific separators based on user settings.
    currency_eur: str = "#,##0.00\ \u20AC"
    integer: str = "#,##0"
    percentage: str = "0.00%"


class GoBDExcelGenerator:
    """Foundational Excel generator for Phase 4 financial reporting.

    This class encapsulates workbook creation, basic properties, common
    styles, and deterministic filename generation. It deliberately avoids
    sheet creation or data fetching to keep Day 11 scope minimal.
    """

    def __init__(
        self,
        organization_name: Optional[str] = None,
        report_title: Optional[str] = None,
        generated_by: Optional[str] = None,
    ) -> None:
        # Initialize workbook and basic document properties
        self._wb = Workbook()
        self._wb.properties.creator = generated_by or "NGO Automation"
        self._wb.properties.title = report_title or "Financial Report"
        self._wb.properties.last_modified_by = generated_by or "NGO Automation"
        self._wb.properties.company = organization_name or "Organization"

        # Formats and styles (German-friendly)
        self.formats = GermanExcelFormats()
        self.styles = self._create_named_styles()

        # Ensure the default sheet is clearly named for later replacement
        self._wb.active.title = "Init"

    # --------- Public API (Day 11 scope) ---------------------------------
    def get_workbook(self) -> Workbook:
        """Return the underlying openpyxl Workbook instance.

        Consumers can add sheets and populate data in subsequent steps.
        """

        return self._wb

    def build_filename(
        self,
        organization_name: str,
        start: date,
        end: date,
        suffix: str = "report",
    ) -> str:
        """Deterministic filename for the exported workbook.

        Example: "NGO_Name_2025-01-01_to_2025-01-31_report.xlsx"
        Unsafe characters are stripped; spaces become underscores.
        """

        safe_org = self._sanitize_filename(organization_name)
        period = f"{start.isoformat()}_to_{end.isoformat()}"
        return f"{safe_org}_{period}_{suffix}.xlsx"

    def build_report_metadata(
        self,
        organization_name: str,
        start: date,
        end: date,
        generated_by: Optional[str] = None,
    ) -> dict:
        """Basic metadata dictionary to accompany the workbook.

        Returns minimal, non-contractual context fields for headers/footers.
        """

        return {
            "organization_name": organization_name,
            "report_period": f"{start.isoformat()} to {end.isoformat()}",
            "generated_at": datetime.utcnow().isoformat(timespec="seconds"),
            "generated_by": generated_by or "System",
        }

    # --------- Internal helpers ------------------------------------------
    def _create_named_styles(self) -> dict[str, NamedStyle]:
        """Create and register commonly used named styles.

        Styles are added once to the workbook's named styles collection and
        reused by later sheet-building steps.
        """

        styles: dict[str, NamedStyle] = {}

        # Header style: bold, center, light fill
        header = NamedStyle(name="header")
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal="center", vertical="center")
        header.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        styles["header"] = header

        # Currency style: right aligned with Euro format
        currency = NamedStyle(name="currency_eur")
        currency.font = Font(bold=False)
        currency.alignment = Alignment(horizontal="right", vertical="center")
        currency.number_format = self.formats.currency_eur
        styles["currency_eur"] = currency

        # Date style: center aligned German date
        date_style = NamedStyle(name="date_de")
        date_style.alignment = Alignment(horizontal="center", vertical="center")
        date_style.number_format = self.formats.date
        styles["date_de"] = date_style

        # Percentage style
        pct = NamedStyle(name="percentage")
        pct.alignment = Alignment(horizontal="right", vertical="center")
        pct.number_format = self.formats.percentage
        styles["percentage"] = pct

        # Register styles (ignore duplicates if already present)
        for style in styles.values():
            try:
                self._wb.add_named_style(style)
            except ValueError:
                # Named style already exists; continue
                pass

        return styles

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        """Sanitize organization name for safe file naming.

        - Strip leading/trailing whitespace
        - Replace spaces with underscores
        - Remove characters not suitable for filenames
        """

        cleaned = name.strip().replace(" ", "_")
        unsafe = "\/:*?\"<>|"
        for ch in unsafe:
            cleaned = cleaned.replace(ch, "")
        return cleaned

# --------- Sheet name constants -------------------------------------------
SHEET_SUMMARY = "Summary"
SHEET_TRANSACTIONS = "Transactions"
SHEET_PROJECTS = "Projects"
SHEET_FEES = "Fees"
SHEET_EVENTS = "Events"
SHEET_AUDIT = "Audit"

# --------- Default headers per sheet (aligned with Phase 4 plan) -----------
DEFAULT_HEADERS_SUMMARY = [
    "Organization",
    "Report Period",
    "Total Revenue (EUR)",
    "Total Expenses (EUR)",
    "Net Position (EUR)",
    "VAT 19% (EUR)",
    "VAT 7% (EUR)",
    "VAT 0% (EUR)",
    "Duplicate Count",
    "Generated At",
]

DEFAULT_HEADERS_TRANSACTIONS = [
    "Date",
    "Vendor",
    "Amount (Gross EUR)",
    "VAT Rate",
    "VAT Amount (EUR)",
    "Net Amount (EUR)",
    "Type",  # expense | revenue
    "Category",
    "Payment Method",
    "Project",
    "Source Type",
    "Transaction Hash",
    "Notes",
    "Running Balance (EUR)",
]

DEFAULT_HEADERS_PROJECTS = [
    "Project ID",
    "Project Name",
    "Transaction Count",
    "Total Amount (EUR)",
    "VAT Total (EUR)",
    "Budget Allocated (EUR)",
    "Budget Remaining (EUR)",
]

DEFAULT_HEADERS_FEES = [
    "Date",
    "Vendor",
    "Amount (EUR)",
    "Category",  # Fees/Services
    "VAT Amount (EUR)",
    "Project",
    "Notes",
]

DEFAULT_HEADERS_EVENTS = [
    "Event Name",
    "Date",
    "Attendees",
    "Cost per Person (EUR)",
    "Total Cost (EUR)",
    "Project",
    "Notes",
]

DEFAULT_HEADERS_AUDIT = [
    "Transaction ID",
    "Transaction Hash",
    "Source Type",
    "Document ID",
    "Created At",
    "Updated At",
    "Is Duplicate",
    "Duplicate Of",
]

# --------- Public API (Day 12: sheet scaffolds with headers) ---------------
def _apply_header(ws: Worksheet, headers: list[str], header_style_name: str = "header") -> None:
    """Write header row (row 1) and apply the named header style."""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.style = header_style_name


def _freeze_header(ws: Worksheet) -> None:
    """Freeze top row so data scrolls beneath the header."""
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, headers: list[str]) -> None:
    """Autosize columns based on header lengths (simple heuristic)."""
    for col_idx, title in enumerate(headers, start=1):
        width = max(len(str(title)), 10) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width


class GoBDExcelGenerator(GoBDExcelGenerator):
    """Extend generator with sheet creation convenience using defaults."""

    def create_summary_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_SUMMARY)
        _apply_header(ws, DEFAULT_HEADERS_SUMMARY)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_SUMMARY)
        return self._wb

    def create_transactions_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_TRANSACTIONS)
        _apply_header(ws, DEFAULT_HEADERS_TRANSACTIONS)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_TRANSACTIONS)
        return self._wb

    def create_projects_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_PROJECTS)
        _apply_header(ws, DEFAULT_HEADERS_PROJECTS)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_PROJECTS)
        return self._wb

    def create_fees_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_FEES)
        _apply_header(ws, DEFAULT_HEADERS_FEES)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_FEES)
        return self._wb

    def create_events_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_EVENTS)
        _apply_header(ws, DEFAULT_HEADERS_EVENTS)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_EVENTS)
        return self._wb

    def create_audit_sheet_default(self) -> Workbook:
        ws = self._wb.create_sheet(title=SHEET_AUDIT)
        _apply_header(ws, DEFAULT_HEADERS_AUDIT)
        _freeze_header(ws)
        _autosize(ws, DEFAULT_HEADERS_AUDIT)
        return self._wb

    # --------- Convenience: build all sheets & save sample -----------------
    def create_all_default_sheets(self) -> Workbook:
        """Create all six default sheets. Removes the initial placeholder sheet.

        Returns the workbook for chaining or saving.
        """
        # Remove placeholder 'Init' if it exists and is empty
        if "Init" in [ws.title for ws in self._wb.worksheets]:
            try:
                self._wb.remove(self._wb["Init"])
            except Exception:
                # Ignore if removal fails; not critical for smoke test
                pass

        self.create_summary_sheet_default()
        self.create_transactions_sheet_default()
        self.create_projects_sheet_default()
        self.create_fees_sheet_default()
        self.create_events_sheet_default()
        self.create_audit_sheet_default()
        return self._wb

    def save_sample_default_workbook(
        self,
        organization_name: str,
        start: date,
        end: date,
        output_dir: str = "test_results",
    ) -> str:
        """Generate all default sheets and save a sample XLSX to output_dir.

        Returns the absolute path to the saved file.
        """
        self.create_all_default_sheets()
        filename = self.build_filename(organization_name, start, end, suffix="excel-smoke")
        if not os.path.isdir(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        self._wb.save(output_path)
        return os.path.abspath(output_path)

    # --------- Day 13-14: data mapping & formulas (no external calls) -----
    def add_transactions(self, rows: Sequence[dict]) -> None:
        """Append transaction rows and compute running balance + VAT/net if missing.

        Expected keys per row (string/number types accepted):
        - date, vendor, amount_gross, vat_rate, vat_amount (optional),
          net_amount (optional), type (expense|revenue), category,
          payment_method, project, source_type, transaction_hash, notes
        """

        ws = self._require_sheet(SHEET_TRANSACTIONS)
        amount_col = self._column_index("Amount (Gross EUR)")
        vat_rate_col = self._column_index("VAT Rate")
        vat_amount_col = self._column_index("VAT Amount (EUR)")
        net_amount_col = self._column_index("Net Amount (EUR)")
        type_col = self._column_index("Type")
        running_col = self._column_index("Running Balance (EUR)")

        start_row = ws.max_row + 1
        for idx, row in enumerate(rows):
            excel_row = start_row + idx
            amount = self._to_decimal(row.get("amount_gross"))
            vat_rate = self._to_decimal(row.get("vat_rate", Decimal("0")))
            vat_amount, net_amount = self._compute_vat(amount, vat_rate, row.get("vat_amount"), row.get("net_amount"))

            # Write cells in header order
            values = [
                row.get("date"),
                row.get("vendor"),
                amount,
                vat_rate,
                vat_amount,
                net_amount,
                row.get("type"),
                row.get("category"),
                row.get("payment_method"),
                row.get("project"),
                row.get("source_type"),
                row.get("transaction_hash"),
                row.get("notes"),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                if col_idx == amount_col or col_idx in (vat_amount_col, net_amount_col):
                    cell.style = "currency_eur"
                elif col_idx == vat_rate_col:
                    cell.number_format = self.formats.percentage
                elif col_idx == 1:  # Date column
                    cell.style = "date_de"

            # Running balance formula: previous balance + signed amount
            self._apply_running_balance(ws, excel_row, amount_col, running_col, type_col)

    def add_projects(self, rows: Sequence[dict]) -> None:
        ws = self._require_sheet(SHEET_PROJECTS)
        start_row = ws.max_row + 1
        for idx, row in enumerate(rows):
            excel_row = start_row + idx
            values = [
                row.get("project_id"),
                row.get("project_name"),
                row.get("transaction_count"),
                self._to_decimal(row.get("total_amount")),
                self._to_decimal(row.get("vat_total")),
                self._to_decimal(row.get("budget_allocated")),
                self._to_decimal(row.get("budget_remaining")),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                if col_idx >= 4:
                    cell.style = "currency_eur"

    def add_fees(self, rows: Sequence[dict]) -> None:
        ws = self._require_sheet(SHEET_FEES)
        start_row = ws.max_row + 1
        for idx, row in enumerate(rows):
            excel_row = start_row + idx
            values = [
                row.get("date"),
                row.get("vendor"),
                self._to_decimal(row.get("amount")),
                row.get("category"),
                self._to_decimal(row.get("vat_amount")),
                row.get("project"),
                row.get("notes"),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                if col_idx in (3, 5):
                    cell.style = "currency_eur"
                elif col_idx == 1:
                    cell.style = "date_de"

    def add_events(self, rows: Sequence[dict]) -> None:
        ws = self._require_sheet(SHEET_EVENTS)
        start_row = ws.max_row + 1
        for idx, row in enumerate(rows):
            excel_row = start_row + idx
            values = [
                row.get("event_name"),
                row.get("date"),
                row.get("attendees"),
                self._to_decimal(row.get("cost_per_person")),
                self._to_decimal(row.get("total_cost")),
                row.get("project"),
                row.get("notes"),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                if col_idx in (4, 5):
                    cell.style = "currency_eur"
                elif col_idx == 2:
                    cell.style = "date_de"

    def add_audit_entries(self, rows: Sequence[dict]) -> None:
        ws = self._require_sheet(SHEET_AUDIT)
        start_row = ws.max_row + 1
        for idx, row in enumerate(rows):
            excel_row = start_row + idx
            values = [
                row.get("transaction_id"),
                row.get("transaction_hash"),
                row.get("source_type"),
                row.get("document_id"),
                row.get("created_at"),
                row.get("updated_at"),
                row.get("is_duplicate"),
                row.get("duplicate_of"),
            ]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=excel_row, column=col_idx, value=val)

    # --------- Internal helpers for mapping --------------------------------
    def _require_sheet(self, title: str) -> Worksheet:
        if title not in [ws.title for ws in self._wb.worksheets]:
            raise ValueError(f"Sheet '{title}' not found. Ensure it is created first.")
        return self._wb[title]

    @staticmethod
    def _to_decimal(value) -> Decimal:
        if value is None:
            return Decimal("0")
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value))
        except Exception:
            return Decimal("0")

    def _compute_vat(
        self,
        gross: Decimal,
        vat_rate: Decimal,
        vat_amount: Optional[Decimal],
        net_amount: Optional[Decimal],
    ) -> tuple[Decimal, Decimal]:
        """Compute VAT/net with simple validation; fallback to arithmetic if missing."""

        rate = vat_rate or Decimal("0")
        if vat_amount is not None and net_amount is not None:
            return (self._to_decimal(vat_amount), self._to_decimal(net_amount))

        # net = gross / (1 + rate); vat = gross - net
        denominator = Decimal("1") + rate
        net_calc = (gross / denominator).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if denominator != 0 else gross
        vat_calc = (gross - net_calc).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return vat_calc, net_calc

    def _apply_running_balance(
        self,
        ws: Worksheet,
        excel_row: int,
        amount_col: int,
        running_col: int,
        type_col: int,
    ) -> None:
        """Apply running balance formula: prev balance + signed amount."""

        amount_cell = ws.cell(row=excel_row, column=amount_col)
        type_cell = ws.cell(row=excel_row, column=type_col)
        running_cell = ws.cell(row=excel_row, column=running_col)

        # Determine sign based on type
        # expense -> negative; revenue -> positive
        sign = -1
        if str(type_cell.value).lower().strip() == "revenue":
            sign = 1

        signed_amount_ref = f"{get_column_letter(amount_col)}{excel_row}"
        signed_expr = f"{sign}*{signed_amount_ref}"

        if excel_row == 2:
            # First data row after header
            running_cell.value = f"={signed_expr}"
        else:
            prev = excel_row - 1
            prev_balance_ref = f"{get_column_letter(running_col)}{prev}"
            running_cell.value = f"={prev_balance_ref}+{signed_expr}"
        running_cell.style = "currency_eur"

    def _column_index(self, header_title: str) -> int:
        """Find column index by header text on the Transactions sheet headers list."""
        try:
            return DEFAULT_HEADERS_TRANSACTIONS.index(header_title) + 1
        except ValueError:
            raise ValueError(f"Header '{header_title}' not found in transactions headers")
